"""Tests for B3 import with XLSX support."""
import pytest
from app.services.b3_import_service import parse_b3_csv, _is_xlsx


def test_is_xlsx_detects_xlsx_files():
    """Test that _is_xlsx correctly identifies XLSX files by magic bytes."""
    # XLSX files start with PK (ZIP signature)
    xlsx_magic = b'PK\x03\x04'
    assert _is_xlsx(xlsx_magic + b'some other data')
    assert _is_xlsx(b'PK\x00\x00\x00')
    
    # CSV files don't start with PK
    csv_content = b'Movimentacao;Data;Produto;Quantidade'
    assert not _is_xlsx(csv_content)
    
    # Empty
    assert not _is_xlsx(b'')


def test_parse_b3_csv_with_xlsx_conversion():
    """Test that parse_b3_csv can accept XLSX files (with openpyxl installed).
    
    This test creates a minimal XLSX file structure and verifies that it's
    converted to CSV and parsed correctly.
    """
    try:
        import openpyxl
        from openpyxl import Workbook
        import io
    except ImportError:
        pytest.skip("openpyxl not available")
    
    # Create a minimal XLSX workbook in memory
    wb = Workbook()
    ws = wb.active
    
    # Header row
    ws.append([
        'Entrada/Saída', 'Data', 'Movimentação', 'Produto', 'Instituição',
        'Quantidade', 'Preço unitário', 'Valor da Operação'
    ])
    
    # Sample buy row
    ws.append([
        'C', '12/03/2024', 'Compra', 'PETR4 - PETROBRAS PN', 'B3',
        '100', '30.00', '3000.00'
    ])
    
    # Save to bytes
    excel_bytes = io.BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)
    content = excel_bytes.read()
    
    # Verify it's detected as XLSX
    assert _is_xlsx(content)
    
    # Parse it - should convert XLSX to CSV and parse successfully
    result = parse_b3_csv(content)
    
    assert len(result.rows) == 1
    # _extract_ticker always appends ".SA" (B3's export is exclusively
    # Brazilian securities, and the rest of Securo stores BR tickers with
    # that suffix) — see the regression this fixed in b3_import_service.
    assert result.rows[0].ticker == 'PETR4.SA'
    assert result.rows[0].kind == 'buy'
    assert result.rows[0].quantity == 100
    
    
def test_parse_b3_csv_with_native_excel_types():
    """Regression test: a real B3 xlsx download has openpyxl hand back
    native `date`/`int`/`float` cells, not display strings — unlike the
    all-string fixture above. This used to fail two ways: (1) the sniffer
    re-parsing the naive ";".join(cells) text misread the delimiter once
    comma-decimal values ("35,50") appeared, collapsing the header into a
    single column and raising a 422 for every real xlsx export; (2) native
    date cells stringified with a trailing " 00:00:00" that _parse_date
    didn't accept, silently dropping every trade row."""
    from datetime import date as _date
    from openpyxl import Workbook
    import io

    wb = Workbook()
    ws = wb.active
    ws.append([
        'Entrada/Saída', 'Data', 'Movimentação', 'Produto', 'Instituição',
        'Quantidade', 'Preço unitário', 'Valor da Operação',
    ])
    ws.append([
        'Credito', _date(2024, 3, 12), 'Compra', 'PETR4 - PETROBRAS PN N2',
        'XP INVESTIMENTOS CCTVM S.A', 100, 35.50, 3550.00,
    ])

    excel_bytes = io.BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)
    content = excel_bytes.read()

    result = parse_b3_csv(content)

    assert result.skipped_count == 0
    assert len(result.rows) == 1
    assert result.rows[0].ticker == 'PETR4.SA'
    assert result.rows[0].kind == 'buy'
    assert result.rows[0].date == _date(2024, 3, 12)


def test_parse_b3_csv_with_malformed_dimension_hint():
    """Regression test: a real B3 xlsx export (investidor.b3.com.br) ships a
    broken <dimension ref="A1"/> tag in the sheet XML — declaring only cell
    A1 even though the sheet has 8 columns and dozens of rows. openpyxl's
    read_only mode trusts that hint to bound iter_rows(), truncating every
    row down to column A alone and collapsing the header into a single
    field ("Entrada/Saída"), which raised a 422 for every real download.
    Rebuilt here by patching a normal workbook's sheet XML the same way."""
    import io
    import re
    import zipfile
    from decimal import Decimal
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append([
        'Entrada/Saída', 'Data', 'Movimentação', 'Produto', 'Instituição',
        'Quantidade', 'Preço unitário', 'Valor da Operação',
    ])
    ws.append([
        'Credito', '30/12/2025', 'Dividendo', 'AZZA3 - AZZAS 2154 S.A.',
        'WARREN DTVM', 18, 1.585, 28.52,
    ])

    good_bytes = io.BytesIO()
    wb.save(good_bytes)
    good_bytes.seek(0)

    with zipfile.ZipFile(good_bytes, 'r') as zin:
        names = zin.namelist()
        contents = {name: zin.read(name) for name in names}

    sheet_path = next(n for n in names if n.startswith('xl/worksheets/sheet'))
    sheet_xml = contents[sheet_path].decode('utf-8')
    broken_xml = re.sub(r'<dimension ref="[^"]*"/>', '<dimension ref="A1"/>', sheet_xml)
    assert broken_xml != sheet_xml, "fixture didn't actually have a <dimension> tag to break"
    contents[sheet_path] = broken_xml.encode('utf-8')

    broken_bytes = io.BytesIO()
    with zipfile.ZipFile(broken_bytes, 'w') as zout:
        for name, data in contents.items():
            zout.writestr(name, data)
    content = broken_bytes.getvalue()

    result = parse_b3_csv(content)

    assert result.skipped_count == 0
    assert len(result.income_rows) == 1
    assert result.income_rows[0].ticker == 'AZZA3.SA'
    assert result.income_rows[0].kind == 'dividendo'
    assert result.income_rows[0].amount == Decimal('28.52')


def test_parse_b3_csv_handles_xlsx_errors():
    """Test that invalid XLSX files raise clear errors."""
    # Invalid XLSX (just ZIP header without valid Excel structure)
    invalid_xlsx = b'PK\x03\x04\x00\x00\x00\x00\x00\x00garbage'
    
    with pytest.raises(ValueError) as exc_info:
        parse_b3_csv(invalid_xlsx)
    
    assert 'XLSX' in str(exc_info.value) or 'Excel' in str(exc_info.value).title()
