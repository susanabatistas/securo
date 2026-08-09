"""Parser for the B3 "Movimentação" CSV export (investidor.b3.com.br →
Extratos → Movimentação). Deterministic, no AI.

Compra/Venda rows become ledger transactions; Dividendo/Juros Sobre Capital
Próprio/Rendimento rows become proventos (AssetIncome, see
app/models/asset_income.py) — separate from the ledger, no unit/cost-basis
impact. Everything else (Bonificação em Ativos, Desdobro, Transferência,
Direitos de Subscrição, ...) is counted and reported back to the caller,
not silently dropped.

The actual weighted-average cost basis is NOT computed here: `apply_b3_rows`
records one AssetTransaction per original CSV row (preserving real trade
dates/prices) and lets the existing ledger (asset_transaction_service,
issue #235) derive units/average_price the same way manual entries do.
`aggregate_for_preview` duplicates that averaging math, but only for
display — it never gets persisted.
"""
from __future__ import annotations

import csv
import io
import re
import unicodedata
import uuid
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Optional

from fastapi import HTTPException
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.asset import Asset
from app.providers.market_price import MarketPriceProvider
from app.providers.tesouro_direto import parse_brl_decimal
from app.schemas.asset import AssetBuyCreate
from app.schemas.asset_income import AssetIncomeCreate
from app.services import asset_income_service, asset_transaction_service

_BUY_KEYWORDS = ("compra",)
_SELL_KEYWORDS = ("venda",)
# Normalized (accent/case-insensitive) substrings for income rows. "juros
# sobre capital proprio" after _normalize() strips the accent on "próprio".
_DIVIDEND_KEYWORDS = ("dividendo",)
_JCP_KEYWORDS = ("juros sobre capital proprio",)
_RENDIMENTO_KEYWORDS = ("rendimento",)


@dataclass(frozen=True)
class B3Row:
    ticker: str
    product: str  # original "Produto" text, e.g. "PETR4 - PETROBRAS PN N2"
    kind: str  # buy | sell
    quantity: Decimal
    price: Decimal  # per unit
    date: date


@dataclass(frozen=True)
class B3IncomeRow:
    ticker: str
    product: str
    kind: str  # dividendo | jcp | rendimento
    amount: Decimal
    date: date


@dataclass(frozen=True)
class B3ParseResult:
    rows: list[B3Row]
    income_rows: list[B3IncomeRow]
    skipped_count: int
    skipped_kinds: dict[str, int] = field(default_factory=dict)


@dataclass(frozen=True)
class B3TickerPreview:
    """Aggregated-by-ticker summary for the preview screen only — not what
    gets persisted (see module docstring)."""

    ticker: str
    buy_quantity: Decimal
    buy_average_price: Decimal
    sell_quantity: Decimal
    sell_average_price: Decimal
    first_date: date
    last_date: date
    row_count: int


def _normalize(value: str) -> str:
    text = unicodedata.normalize("NFKD", value or "")
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", text).strip().casefold()


def _header_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", _normalize(value))


def _clean_decimal_str(value: str) -> str:
    """Strip currency symbols/whitespace some brokers' exports embed in
    numeric fields (seen from Warren: "R$0,10 ") before handing off to
    parse_brl_decimal, which only expects digits/separators. Keeps digits,
    comma, dot, and minus sign; drops everything else (currency symbols,
    NBSP, trailing spaces)."""
    return re.sub(r"[^\d,.\-]", "", value or "")


def _sniff_dialect(text: str):
    try:
        return csv.Sniffer().sniff(text[:4096], delimiters=";,\t")
    except csv.Error:
        return csv.excel


def _is_xlsx(content: bytes) -> bool:
    """XLSX files are ZIP archives — detected by the ZIP local-file-header
    magic bytes ("PK"), same signature the openpyxl/zipfile stack itself
    keys off. A B3 CSV export never starts with those bytes."""
    return bool(content) and content.startswith(b"PK")


def _xlsx_to_delimited_text(content: bytes) -> str:
    """First worksheet of an XLSX file, as ';'-delimited text — reuses the
    exact same column-detection/row-parsing path as a native CSV export
    below (B3's Excel download has the same columns, just a different
    container). Raises ValueError (mentioning XLSX/Excel) for anything that
    isn't a readable workbook, same contract as the CSV path's errors."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover — declared dependency
        raise ValueError("Suporte a arquivos XLSX indisponível neste servidor.") from exc

    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        worksheet = workbook.active
        lines = []
        for row in worksheet.iter_rows(values_only=True):
            cells = ["" if cell is None else str(cell) for cell in row]
            lines.append(";".join(cells))
        return "\n".join(lines)
    except ValueError:
        raise
    except Exception as exc:
        raise ValueError(f"Não foi possível ler o arquivo XLSX: {exc}") from exc


def _parse_date(value: str) -> date:
    raw = (value or "").strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            pass
    raise ValueError(f"invalid date: {value}")


def _extract_ticker(product: str) -> str:
    """B3's "Produto" column is "TICKER - Company Name", e.g.
    "PETR4 - PETROBRAS PN N2" -> "PETR4". Falls back to the first
    whitespace-separated token when there's no " - " separator.

    Always appends ".SA" — the B3 Movimentação export is exclusively
    B3-listed (Brazilian) securities, and the rest of Securo (the yfinance-
    backed market_price flow, buy_into_holding, etc.) always stores BR
    tickers with that suffix. Without it, every B3-imported row would land
    on a bare-ticker asset instead of consolidating into (or matching, for
    proventos) the existing "TICKER.SA" holding — confirmed with a real
    user CSV where 23/23 proventos failed to match for exactly this reason.
    """
    text = (product or "").strip()
    base = text.split(" - ", 1)[0].strip().upper() if " - " in text else (text.split()[0] if text else "").upper()
    if not base or base.endswith(".SA"):
        return base
    return f"{base}.SA"


_MOVEMENT_COL = ("movimentacao", "tipodemovimentacao", "tipomovimentacao")
_DATE_COL = ("data", "dataregistro", "data_do_negocio")
_PRODUCT_COL = ("produto", "ativo")
_QUANTITY_COL = ("quantidade",)
_UNIT_PRICE_COL = ("precounitario", "preco", "precomedio")
_OPERATION_VALUE_COL = ("valordaoperacao", "valor", "valorbruto")


def parse_b3_csv(content: bytes) -> B3ParseResult:
    """Parse a B3 "Movimentação" export in CSV format. Raises ValueError with a clear
    message (available columns included) if the essential columns can't be
    recognized — never silently returns a bogus/partial result.

    B3's "Movimentação" export can be downloaded as XLSX instead of CSV —
    detected by magic bytes and converted to the same delimited-text shape
    before the rest of this function (column detection, row parsing) runs
    unchanged; XLSX is just a different container for the same columns.
    """
    if _is_xlsx(content):
        text = _xlsx_to_delimited_text(content)
    else:
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = content.decode("latin1")

    if not text.strip():
        raise ValueError("empty file")

    dialect = _sniff_dialect(text)
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    if not reader.fieldnames:
        raise ValueError("could not read a header row from this file")

    header_map = {_header_key(name): name for name in reader.fieldnames if name}

    def find(candidates: tuple[str, ...]) -> str | None:
        for c in candidates:
            if c in header_map:
                return header_map[c]
        return None

    movement_col = find(_MOVEMENT_COL)
    date_col = find(_DATE_COL)
    product_col = find(_PRODUCT_COL)
    quantity_col = find(_QUANTITY_COL)
    unit_price_col = find(_UNIT_PRICE_COL)
    operation_value_col = find(_OPERATION_VALUE_COL)

    missing_essentials = not (movement_col and date_col and product_col and quantity_col)
    if missing_essentials or not (unit_price_col or operation_value_col):
        raise ValueError(
            "Formato de CSV não reconhecido como extrato de Movimentação da B3. "
            f"Colunas encontradas: {', '.join(reader.fieldnames)}"
        )

    rows: list[B3Row] = []
    income_rows: list[B3IncomeRow] = []
    skipped_count = 0
    skipped_kinds: dict[str, int] = {}

    for raw_row in reader:
        movement_raw = (raw_row.get(movement_col) or "").strip()
        movement_norm = _normalize(movement_raw)
        product = raw_row.get(product_col) or ""
        ticker = _extract_ticker(product)

        if any(k in movement_norm for k in _BUY_KEYWORDS):
            trade_kind: Optional[str] = "buy"
            income_kind = None
        elif any(k in movement_norm for k in _SELL_KEYWORDS):
            trade_kind = "sell"
            income_kind = None
        elif any(k in movement_norm for k in _DIVIDEND_KEYWORDS):
            trade_kind, income_kind = None, "dividendo"
        elif any(k in movement_norm for k in _JCP_KEYWORDS):
            trade_kind, income_kind = None, "jcp"
        elif any(k in movement_norm for k in _RENDIMENTO_KEYWORDS):
            trade_kind, income_kind = None, "rendimento"
        else:
            skipped_count += 1
            label = movement_raw or "(vazio)"
            skipped_kinds[label] = skipped_kinds.get(label, 0) + 1
            continue

        try:
            row_date = _parse_date(raw_row.get(date_col) or "")
            if not ticker:
                raise ValueError("missing ticker")

            if trade_kind is not None:
                quantity = parse_brl_decimal(_clean_decimal_str(raw_row.get(quantity_col) or ""))
                if quantity <= 0:
                    raise ValueError("non-positive quantity")
                if unit_price_col and (raw_row.get(unit_price_col) or "").strip():
                    price = parse_brl_decimal(_clean_decimal_str(raw_row[unit_price_col]))
                else:
                    operation_value = parse_brl_decimal(_clean_decimal_str(raw_row.get(operation_value_col) or ""))
                    price = (operation_value / quantity).quantize(Decimal("0.000001"))
                rows.append(
                    B3Row(ticker=ticker, product=product, kind=trade_kind, quantity=quantity, price=price, date=row_date)
                )
            else:
                # Income rows carry their total in "Valor da Operação" —
                # there's no per-unit price to derive here.
                amount = parse_brl_decimal(_clean_decimal_str(raw_row.get(operation_value_col) or ""))
                if amount <= 0:
                    raise ValueError("non-positive amount")
                income_rows.append(
                    B3IncomeRow(ticker=ticker, product=product, kind=income_kind, amount=amount, date=row_date)
                )
        except (ValueError, InvalidOperation, ZeroDivisionError):
            # A single malformed data row shouldn't sink the whole import —
            # counted as skipped, same as an unrecognized movement type.
            skipped_count += 1
            skipped_kinds["(linha inválida)"] = skipped_kinds.get("(linha inválida)", 0) + 1
            continue

    return B3ParseResult(rows=rows, income_rows=income_rows, skipped_count=skipped_count, skipped_kinds=skipped_kinds)


def aggregate_for_preview(rows: list[B3Row]) -> list[B3TickerPreview]:
    """Weighted-average summary per ticker, for the preview screen only."""
    by_ticker: dict[str, list[B3Row]] = {}
    for row in rows:
        by_ticker.setdefault(row.ticker, []).append(row)

    previews: list[B3TickerPreview] = []
    for ticker, ticker_rows in sorted(by_ticker.items()):
        buy_qty = Decimal("0")
        buy_cost = Decimal("0")
        sell_qty = Decimal("0")
        sell_cost = Decimal("0")
        for r in ticker_rows:
            if r.kind == "buy":
                buy_qty += r.quantity
                buy_cost += r.quantity * r.price
            else:
                sell_qty += r.quantity
                sell_cost += r.quantity * r.price
        dates = [r.date for r in ticker_rows]
        previews.append(
            B3TickerPreview(
                ticker=ticker,
                buy_quantity=buy_qty,
                buy_average_price=(buy_cost / buy_qty).quantize(Decimal("0.000001")) if buy_qty > 0 else Decimal("0"),
                sell_quantity=sell_qty,
                sell_average_price=(sell_cost / sell_qty).quantize(Decimal("0.000001")) if sell_qty > 0 else Decimal("0"),
                first_date=min(dates),
                last_date=max(dates),
                row_count=len(ticker_rows),
            )
        )
    return previews


@dataclass(frozen=True)
class B3ApplyError:
    ticker: str
    kind: str
    date: date
    reason: str


@dataclass(frozen=True)
class B3ApplyResult:
    applied_count: int
    income_applied_count: int
    errors: list[B3ApplyError]


async def _find_asset_by_ticker(
    session: AsyncSession, workspace_id: uuid.UUID, ticker: str, group_id: Optional[uuid.UUID]
) -> Optional[Asset]:
    # No sell_date/is_archived filter — a dividend can be logged for a
    # position already sold (same rule as asset_income_service._load_asset).
    result = await session.execute(
        select(Asset).where(
            Asset.workspace_id == workspace_id,
            Asset.ticker == ticker,
            Asset.valuation_method == "market_price",
            Asset.group_id == group_id,
        )
    )
    return result.scalar_one_or_none()


async def _resolve_group_id(
    session: AsyncSession, workspace_id: uuid.UUID, ticker: str, requested_group_id: Optional[uuid.UUID]
) -> Optional[uuid.UUID]:
    """An explicit wallet always wins. Otherwise, auto-detect: if this
    ticker already exists in some wallet, use that wallet — a B3 import
    doesn't ask "which wallet?" per row, so without this every real
    portfolio (assets are almost always organized into wallets) would
    silently target "ungrouped" and never match anything. Confirmed with a
    real user whose entire portfolio lives in 4 wallets: every one of their
    23 proventos failed with "sem posição encontrada" until this existed.

    If the same ticker exists in more than one wallet, this picks one
    arbitrarily (B3's export doesn't carry a wallet identifier Securo can
    map to — only a broker name, which isn't matched here) — a real but
    rare edge case, left for the user to fix by moving/re-entering by hand.
    """
    if requested_group_id is not None:
        return requested_group_id
    result = await session.execute(
        select(Asset.group_id)
        .where(
            Asset.workspace_id == workspace_id,
            Asset.ticker == ticker,
            Asset.valuation_method == "market_price",
        )
        .limit(1)
    )
    row = result.first()
    return row[0] if row else None


async def apply_b3_rows(
    session: AsyncSession,
    workspace_id: uuid.UUID,
    user_id: uuid.UUID,
    rows: list[B3Row],
    *,
    income_rows: Optional[list[B3IncomeRow]] = None,
    group_id: Optional[uuid.UUID] = None,
    market_provider: Optional[MarketPriceProvider] = None,
) -> B3ApplyResult:
    """Apply parsed B3 rows — buy/sell as ledger transactions (one
    AssetTransaction per row, see module docstring), then proventos as
    AssetIncome. Trades are applied first, in date order, so a holding
    created by a buy in this same import is visible to a later dividend row
    for the same ticker. A row that fails (e.g. a sell/dividend with no
    matching holding, or an oversell) is recorded in `errors` and does NOT
    abort the rest of the import.
    """
    applied = 0
    income_applied = 0
    errors: list[B3ApplyError] = []

    for row in sorted(rows, key=lambda r: r.date):
        try:
            row_group_id = await _resolve_group_id(session, workspace_id, row.ticker, group_id)
            if row.kind == "buy":
                await asset_transaction_service.buy_into_holding(
                    session, workspace_id, user_id,
                    AssetBuyCreate(
                        ticker=row.ticker, quantity=row.quantity, price=row.price,
                        date=row.date, group_id=row_group_id,
                    ),
                    market_provider=market_provider,
                    source="import",
                )
            else:
                sold = await asset_transaction_service.sell_from_holding(
                    session, workspace_id, row.ticker, row_group_id,
                    row.quantity, row.price, row.date, source="import",
                )
                if sold is None:
                    errors.append(
                        B3ApplyError(
                            ticker=row.ticker, kind=row.kind, date=row.date,
                            reason="venda sem posição encontrada",
                        )
                    )
                    continue
            applied += 1
        except HTTPException as e:
            errors.append(
                B3ApplyError(ticker=row.ticker, kind=row.kind, date=row.date, reason=str(e.detail))
            )

    for income_row in sorted(income_rows or [], key=lambda r: r.date):
        income_group_id = await _resolve_group_id(session, workspace_id, income_row.ticker, group_id)
        asset = await _find_asset_by_ticker(session, workspace_id, income_row.ticker, income_group_id)
        if asset is None:
            errors.append(
                B3ApplyError(
                    ticker=income_row.ticker, kind=income_row.kind, date=income_row.date,
                    reason="provento sem posição encontrada",
                )
            )
            continue
        result = await asset_income_service.add_income(
            session, asset.id, workspace_id,
            AssetIncomeCreate(kind=income_row.kind, amount=income_row.amount, date=income_row.date),
            source="import",
        )
        if result is None:  # pragma: no cover — asset was just resolved above
            continue
        income_applied += 1

    return B3ApplyResult(applied_count=applied, income_applied_count=income_applied, errors=errors)
